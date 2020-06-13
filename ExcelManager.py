import openpyxl
import os


def delet_excel(excel_name, base_address=""):
    try:
        if len(excel_name) == 0:
            return
        if len(base_address) != 0:
            if base_address[-1] != "\\":
                base_address += "\\"
        address = base_address + excel_name + ".xlsx"
        if os.path.exists(address):
            os.remove(address)
        else:
            print("The file does not exist")
    except Exception as e:
        print("[EXCET] delete excel ", e)


def get_create_excel(excel_name, base_address=""):
    if len(excel_name) == 0:
        return
    if len(base_address) != 0:
        if base_address[-1] != "\\":
            base_address += "\\"
    address = base_address + excel_name + ".xlsx"
    try:
        workbook = openpyxl.load_workbook(address)
    except Exception as e:
        workbook = openpyxl.Workbook(address)
        print(excel_name, ".xlsx", " created")

    return workbook
    # workbook.close()


def create_sheet(excel_name, sheet_name, columns_name, base_address=""):
    try:
        if len(excel_name) == 0:
            return None
        if len(base_address) != 0:
            if base_address[-1] != "\\":
                base_address += "\\"
        address = base_address + excel_name + ".xlsx"
        workbook = get_create_excel(base_address=base_address, excel_name=excel_name)
        if sheet_name in workbook.sheetnames:
            print("this sheet is existed : ", sheet_name)
            return
        else:
            worksheet = workbook.create_sheet(sheet_name)
            worksheet.append(columns_name)
            print(sheet_name, "  sheet in ", excel_name, " excel  created")
            workbook.save(address)
            workbook.close()
    except Exception as e:
        print("create sheet exception ", e)


def add_row(excel_name, sheet_name, data, base_address="", foreign_key=None):
    if len(excel_name) == 0:
        return
    if len(base_address) != 0:
        if base_address[-1] != "\\":
            base_address += "\\"
    address = base_address + excel_name + ".xlsx"
    try:
        workbook = openpyxl.load_workbook(address)
        sheet = workbook.get_sheet_by_name(name=sheet_name)
        if foreign_key is not None:
            index = -1
            items = sheet.iter_rows()
            items = list(items)
            row = list(items)[-1]
            if str(row[foreign_key[1]].value).strip() == foreign_key[0]:
                index = len(list(items))
            if index > -1:
                sheet.delete_rows(index)
        sheet.append(data)
        workbook.save(address)
        workbook.close()
    except Exception as e:
        print(e)
    print(sheet_name, " sheet in ", excel_name, "excel is updated")


def add_rows(excel_name, sheet_name, datas, base_address="", foreign_keys=None):
    if len(excel_name) == 0:
        return
    if len(base_address) != 0:
        if base_address[-1] != "\\":
            base_address += "\\"
    address = base_address + excel_name + ".xlsx"
    try:
        workbook = openpyxl.load_workbook(address)
        sheet = workbook.get_sheet_by_name(name=sheet_name)
        if foreign_keys is not None:
            for foreign_key in foreign_keys:
                items = sheet.iter_rows()
                items = list(items)
                for index in range(len(items) - 1, -1, -1):
                    row = list(items)[index]
                    if str(row[foreign_key[1]].value).strip() == foreign_key[0]:
                        sheet.delete_rows(index + 1)

        for data in datas:
            sheet.append(data)
        workbook.save(address)
        workbook.close()
    except Exception as e:
        print("EXCEPT ", e)
    print(sheet_name, " sheet in ", excel_name, "excel is updated")


def read_rows(excel_name, sheet_name, base_address=""):
    if len(excel_name) == 0:
        return
    if len(base_address) != 0:
        if base_address[-1] != "\\":
            base_address += "\\"
    address = base_address + excel_name + ".xlsx"
    try:
        workbook = openpyxl.load_workbook(address)
        sheet = workbook.get_sheet_by_name(name=sheet_name)
        rows = list(sheet.iter_rows())
        informations = []
        for row in rows:
            informations.append([])
            for col in row:
                informations[-1].append(col.value)

        return informations

    except Exception as e:
        raise e


if __name__ == '__main__':
    datas = read_rows(excel_name="test", sheet_name="Online Retail")
